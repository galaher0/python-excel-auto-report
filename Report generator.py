import openpyxl
import os
from datetime import datetime
import tkcalendar as tkc
import tkinter as tk


class Program(tk.Tk):

	def __init__(self, *args, **kwargs):
		super().__init__(*args, **kwargs)
		self.title('авто леха')
		self.list_files = [''] + [f for f in os.listdir('.') if f.endswith(".xlsx")]
		tk.Label(self, text="Файл с отгрузками (заказами)", padx=5, pady=5).grid(sticky="W", row=0, column=0, padx=5, pady=5)
		tk.Label(self, text="Матрица Леруа", padx=5, pady=5).grid(sticky="W", row=1, column=0, padx=5, pady=5)
		self.despatch_file_var = tk.StringVar()
		self.leroy_matrix_file_var = tk.StringVar()
		tk.OptionMenu(self, self.despatch_file_var, *self.list_files).grid(row=0, column=1, padx=5, pady=5)
		tk.OptionMenu(self, self.leroy_matrix_file_var, *self.list_files).grid(row=1, column=1, padx=5, pady=5)
		tk.Button(self, text='Далее', command=self.check, padx=5, pady=5).grid(row=3, column=0, columnspan=2, padx=5, pady=5)

	def check(self):
		if not self.despatch_file_var.get() or not self.leroy_matrix_file_var.get():
			tk.Label(self, text="Не выбраны входные файлы", padx=5, pady=5).grid(row=2, column=0, columnspan=2, padx=5, pady=5)
		else:
			if self.grid_slaves(row=2, column=0):
				self.grid_slaves(row=2, column=0)[0].grid_forget()

			self.despatches_wb = openpyxl.load_workbook(self.despatch_file_var.get())
			self.wsr = self.despatches_wb.worksheets[0]
			working_columns = utils.find_not_hidden_cols(self.wsr)
			self.appr_columns = {c: tk.BooleanVar(value=True) for c in working_columns}

			appr_frame = tk.Frame(self)
			tk.Label(appr_frame, text="Подтвердите выбор столбцов:", padx=5, pady=5).grid(sticky="W", row=0, column=0, padx=5, pady=5)
			for i, c in enumerate(self.appr_columns):
				tk.Checkbutton(appr_frame, text=c, variable=self.appr_columns[c]).grid(sticky="W", row=i+1, column=0, padx=10)

			tk.Label(appr_frame, text="Выберите временной период:").grid(row=0, column=1, columnspan=2, padx=5, pady=5)
			tk.Label(appr_frame, text="Начало").grid(row=1, column=1, padx=5, pady=5)
			self.start_date = tkc.DateEntry(appr_frame, date_pattern="dd.mm.y", locale='ru_RU')
			self.start_date.grid(row=1, column=2, padx=5, pady=5)
			tk.Label(appr_frame, text="Конец").grid(row=2, column=1, padx=5, pady=5)
			self.end_date = tkc.DateEntry(appr_frame, date_pattern="dd.mm.y", locale='ru_RU')
			self.end_date.grid(row=2, column=2, padx=5, pady=5)
			appr_frame.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

			self.grid_slaves(row=3, column=0)[0].grid_forget()
			tk.Button(self, text='Создать отчет', command=self.create_report, padx=5, pady=5).grid(row=3, column=0, columnspan=2, padx=5, pady=5)

	def create_report(self):
		self.dates = utils.find_valid_dates(self.start_date.get_date(), self.end_date.get_date(), self.wsr)
		row_ranges = [utils.find_range_by_date(i, self.wsr) for i in self.dates]
		headers = [
			'дата консолидации',
			'дата отгрузки на РЦ',
			'магазин',
			'Группа HG/Остальное',
			'ЛМ код',
			'арт. поставщика',
			'наименование',
			'цена руб.',
			'с НДС',
			'кол-во',
			'общая сумма заказа'
		]
		columns_to_write = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
		columns_to_read = [c for c in self.appr_columns if self.appr_columns[c].get()]
		ddmmyyyy = openpyxl.styles.NamedStyle(name='ds', number_format="DD.MM.YYYY")

		matrix = openpyxl.load_workbook(self.leroy_matrix_file_var.get())
		lookup_sheet = 'Hansgrohe'

		report = openpyxl.Workbook()
		wsw = report.active

		for i, col in enumerate(columns_to_write):
			wsw[col+'1'] = headers[i]

		row_counter = 2
		for cr in columns_to_read:
			for i, row_range in enumerate(row_ranges):
				data_rows = utils.return_triples(cr, row_range, self.wsr)
				for valid_row in data_rows[0]:
					wsw['A'+str(row_counter)].value = self.dates[i] # дата
					wsw['A'+str(row_counter)].style = ddmmyyyy # формат ячейки
					wsw['C'+str(row_counter)].value = self.wsr[cr+'1'].value # название магаза
					wsw['D'+str(row_counter)].value = 'Hansgrohe' if utils.is_hansgrohe(valid_row[0]) else 'остальное'
					wsw['E'+str(row_counter)].value = f"=VLOOKUP(F{row_counter},'[{self.leroy_matrix_file_var.get()}]{lookup_sheet}'!$B:$C,2,FALSE)"
					wsw['F'+str(row_counter)].value = valid_row[0]
					wsw['G'+str(row_counter)].value = f"=VLOOKUP(E{row_counter},'[{self.leroy_matrix_file_var.get()}]{lookup_sheet}'!$C:$D,2,FALSE)"
					wsw['H'+str(row_counter)].value = valid_row[2]
					wsw['I'+str(row_counter)].value = f"=H{row_counter}*1.2"
					wsw['J'+str(row_counter)].value = valid_row[1]
					wsw['K'+str(row_counter)].value = f"=I{row_counter}*J{row_counter}"
					row_counter += 1
				for invalid_row in data_rows[1]:
					wsw['A'+str(row_counter)].value = self.dates[i] #  дата
					wsw['A'+str(row_counter)].style = ddmmyyyy 
					wsw['C'+str(row_counter)].value = self.wsr[cr+'1'].value # название магаза
					wsw['D'+str(row_counter)].value = 'Hansgrohe' if utils.is_hansgrohe(invalid_row[0]) else 'остальное'
					wsw['E'+str(row_counter)].value = f"=VLOOKUP(F{row_counter},'[{self.leroy_matrix_file_var.get()}]{lookup_sheet}'!$B:$C,2,FALSE)"
					wsw['F'+str(row_counter)].value = invalid_row[0]
					wsw['G'+str(row_counter)].value = f"=VLOOKUP(E{row_counter},'[{self.leroy_matrix_file_var.get()}]{lookup_sheet}'!$C:$D,2,FALSE)"
					wsw['J'+str(row_counter)].value = invalid_row[1]
					wsw['K'+str(row_counter)].value = f"=I{row_counter}*J{row_counter}"
					row_counter += 1

		report.save('Report.xlsx')
		tk.Label(self, text='Отчет создан, можно закрывать программу').grid(row=4, column=0, columnspan=2, padx=5, pady=5)

	def run(self):
		self.mainloop()



class utils:
	@staticmethod
	def find_not_hidden_cols(worksheet_obj):
		max_col = worksheet_obj.max_column
		working_cols = []
		for col_id in range(2, max_col+1):
			id_as_str = openpyxl.utils.get_column_letter(col_id)
			if not worksheet_obj.column_dimensions[id_as_str].hidden:
				working_cols.append(id_as_str)
		return working_cols

	@staticmethod
	def find_valid_dates(start, end, ws):
		dates = []
		cells_A = next(ws.columns)
		for cell in cells_A:
			if cell.value and isinstance(cell.value, datetime):
				if start <= cell.value.date() <= end:
					dates.append(cell.value)
		return dates

	@staticmethod
	def find_range_by_date(datetime_obj, worksheet_obj):
		cells_A = next(worksheet_obj.columns)
		for cell in cells_A:
			if cell.value == datetime_obj:
				start = cell.row
				break
		end = cells_A[-1].row
		for cell in cells_A[start:]:
			if cell.value:
				end = cell.row - 1
				break
		return (start, end)

	@staticmethod
	def is_vendor_code(value):
		if value:
			if type(value) is str or value >= 100000:
				return True
		return False

	@staticmethod
	def is_amount(value):
		if type(value) is int:
			if value < 100:
				return True
		return False

	@staticmethod
	def is_price(value):
		if value:
			if (type(value) is float or type(value) is int) and 100 < value < 100000:
				return True
			elif type(value) is str:
				value = ''.join(c for c in value if c.isdigit() or c == ',').replace(',','.')
				try:
					value = float(value)
				except:
					return False
				return utils.is_price(value)
		return False

	@staticmethod
	def is_valid_triple(list_):
		return len(list_) == 3 and utils.is_vendor_code(list_[0]) and \
		utils.is_amount(list_[1]) and utils.is_price(list_[2])

	@staticmethod
	def is_hansgrohe(value):
		return type(value) is int and len(str(value)) == 8

	@staticmethod
	def return_triples(column, row_range, worksheet_obj):
		valids = []
		invalids = []
		col = column
		rows = row_range
		ws = worksheet_obj
		
		for i in range(rows[0], rows[1]+1, 3):
			if ws[col+str(i)].value:
				triple = [ws[col+str(i)].value, ws[col+str(i+1)].value, ws[col+str(i+2)].value]
				if utils.is_valid_triple(triple):
					valids.append(triple)
				else:
					invalids.append(triple)
					for j in range(i+1, rows[1]+1):
						if utils.is_vendor_code(ws[col+str(j)].value):
							res = utils.return_triples(column, (j, rows[1]), ws)
							valids = valids + res[0]
							invalids = invalids + res[1]
							break
					break
		return (valids, invalids)



p = Program()
p.run()